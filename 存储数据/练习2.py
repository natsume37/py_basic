from openpyxl import load_workbook

# 打开 Excel 文件
wb = load_workbook('考勤统计.xlsx')
# 选择考勤统计表这张工作表
sheet = wb['考勤统计表']
# 打印出所有工作表名称
print(wb.sheetnames)
# 输出：['考勤统计表']

# 打印出 A1 单元格的值
print(sheet['A1'].value)
# 输出：姓名

# 打印所有单元格的值
for row in sheet.rows:
  for cell in row:
    print(cell.value)